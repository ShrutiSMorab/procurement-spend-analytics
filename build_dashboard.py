"""Builds spend_dashboard.xlsx from spend.db — all KPIs as live Excel formulas."""
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, LineChart, Reference

conn = sqlite3.connect("spend.db")
cur = conn.cursor()

rows = cur.execute("""
    SELECT strftime('%Y-%m', p.po_date), c.category_name, b.bu_name,
           s.supplier_name,
           CASE WHEN ct.contract_id IS NULL THEN 'N' ELSE 'Y' END,
           p.quantity, p.unit_price
    FROM po_lines p
    JOIN categories c ON c.category_id = p.category_id
    JOIN business_units b ON b.bu_id = p.bu_id
    JOIN suppliers s ON s.supplier_id = p.supplier_id
    LEFT JOIN contracts ct ON ct.category_id = p.category_id
         AND ct.supplier_id = p.supplier_id
         AND p.po_date BETWEEN ct.valid_from AND ct.valid_to
    ORDER BY p.po_date
""").fetchall()

cats = [r[0] for r in cur.execute(
    "SELECT category_name FROM categories ORDER BY category_name")]
months = [r[0] for r in cur.execute(
    "SELECT DISTINCT strftime('%Y-%m', po_date) FROM po_lines ORDER BY 1")]
bus = [r[0] for r in cur.execute("SELECT bu_name FROM business_units")]
sups = [r[0] for r in cur.execute("""
    SELECT s.supplier_name FROM po_lines p
    JOIN suppliers s ON s.supplier_id = p.supplier_id
    GROUP BY s.supplier_id ORDER BY SUM(p.quantity*p.unit_price) DESC""")]

wb = Workbook()
BOLD = Font(name="Arial", bold=True)
BASE = Font(name="Arial")
HDR_FILL = PatternFill("solid", start_color="1F4E78")
HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF")
EUR = '€#,##0'
PCT = '0.0%'

def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL

# ---- Data sheet
d = wb.active
d.title = "Data"
d.append(["Month", "Category", "BusinessUnit", "Supplier", "OnContract",
          "Quantity", "UnitPrice", "LineTotal"])
for i, r in enumerate(rows, start=2):
    d.append(list(r) + [f"=F{i}*G{i}"])
style_header(d, 8)
n = len(rows) + 1
for col, w in zip("ABCDEFGH", [10, 22, 18, 22, 11, 9, 10, 12]):
    d.column_dimensions[col].width = w

# ---- Category sheet
cs = wb.create_sheet("ByCategory")
cs.append(["Category", "Spend", "PctOfTotal"])
for i, cat in enumerate(cats, start=2):
    cs.append([cat,
               f'=SUMIF(Data!B:B,A{i},Data!H:H)',
               f'=B{i}/SUM(B$2:B${len(cats)+1})'])
style_header(cs, 3)
for i in range(2, len(cats) + 2):
    cs[f"B{i}"].number_format = EUR
    cs[f"C{i}"].number_format = PCT
cs.column_dimensions["A"].width = 24
cs.column_dimensions["B"].width = 14

# ---- Monthly sheet
ms = wb.create_sheet("ByMonth")
ms.append(["Month", "Spend"])
for i, m in enumerate(months, start=2):
    ms.append([m, f'=SUMIF(Data!A:A,A{i},Data!H:H)'])
style_header(ms, 2)
for i in range(2, len(months) + 2):
    ms[f"B{i}"].number_format = EUR
ms.column_dimensions["B"].width = 14

# ---- Maverick by BU
mb = wb.create_sheet("Maverick")
mb.append(["BusinessUnit", "Spend", "MaverickSpend", "MaverickPct"])
for i, b in enumerate(bus, start=2):
    mb.append([b,
               f'=SUMIF(Data!C:C,A{i},Data!H:H)',
               f'=SUMIFS(Data!H:H,Data!C:C,A{i},Data!E:E,"N")',
               f'=C{i}/B{i}'])
style_header(mb, 4)
for i in range(2, len(bus) + 2):
    mb[f"B{i}"].number_format = EUR
    mb[f"C{i}"].number_format = EUR
    mb[f"D{i}"].number_format = PCT
mb.column_dimensions["A"].width = 18
mb.column_dimensions["B"].width = 14
mb.column_dimensions["C"].width = 14

# ---- Suppliers (pre-sorted by spend desc; values are live formulas)
ss = wb.create_sheet("Suppliers")
ss.append(["Supplier", "Spend", "PctOfTotal", "CumulativePct"])
for i, s in enumerate(sups, start=2):
    ss.append([s,
               f'=SUMIF(Data!D:D,A{i},Data!H:H)',
               f'=B{i}/SUM(B$2:B${len(sups)+1})',
               f'=SUM(C$2:C{i})'])
style_header(ss, 4)
for i in range(2, len(sups) + 2):
    ss[f"B{i}"].number_format = EUR
    ss[f"C{i}"].number_format = PCT
    ss[f"D{i}"].number_format = PCT
ss.column_dimensions["A"].width = 24
ss.column_dimensions["B"].width = 14

# ---- Dashboard
db = wb.create_sheet("Dashboard", 0)
db["A1"] = "Indirect Procurement Spend Dashboard"
db["A1"].font = Font(name="Arial", bold=True, size=16)
db["A2"] = "24 months synthetic spend data — all figures calculated live from the Data sheet"
db["A2"].font = Font(name="Arial", italic=True, size=9)

kpis = [
    ("Total spend", "=SUM(Data!H:H)", EUR),
    ("PO lines", "=COUNTA(Data!A:A)-1", '#,##0'),
    ("Active suppliers", '=COUNTIF(Suppliers!B:B,">0")', '#,##0'),
    ("Maverick spend %", '=SUMIF(Data!E:E,"N",Data!H:H)/SUM(Data!H:H)', PCT),
    ("Suppliers for 80% of spend", '=COUNTIF(Suppliers!D2:D110,"<=0.8")+1', '#,##0'),
]
for j, (label, formula, fmt) in enumerate(kpis):
    col = 1 + j * 2
    lc = db.cell(row=4, column=col, value=label)
    vc = db.cell(row=5, column=col, value=formula)
    lc.font = BOLD
    vc.font = Font(name="Arial", size=13)
    vc.number_format = fmt

for col in "ACEGI":
    db.column_dimensions[col].width = 16

bar = BarChart(); bar.title = "Spend by category (EUR)"; bar.style = 10
bar.add_data(Reference(cs, min_col=2, min_row=1, max_row=len(cats)+1), titles_from_data=True)
bar.set_categories(Reference(cs, min_col=1, min_row=2, max_row=len(cats)+1))
bar.height, bar.width, bar.legend = 8, 16, None
db.add_chart(bar, "A8")

line = LineChart(); line.title = "Monthly spend trend (EUR)"; line.style = 12
line.add_data(Reference(ms, min_col=2, min_row=1, max_row=len(months)+1), titles_from_data=True)
line.set_categories(Reference(ms, min_col=1, min_row=2, max_row=len(months)+1))
line.height, line.width, line.legend = 8, 16, None
db.add_chart(line, "J8")

bar2 = BarChart(); bar2.title = "Maverick spend % by business unit"; bar2.style = 10
bar2.add_data(Reference(mb, min_col=4, min_row=1, max_row=len(bus)+1), titles_from_data=True)
bar2.set_categories(Reference(mb, min_col=1, min_row=2, max_row=len(bus)+1))
bar2.height, bar2.width, bar2.legend = 8, 16, None
db.add_chart(bar2, "A25")

bar3 = BarChart(); bar3.title = "Top 10 suppliers by spend (EUR)"; bar3.style = 10
bar3.add_data(Reference(ss, min_col=2, min_row=1, max_row=11), titles_from_data=True)
bar3.set_categories(Reference(ss, min_col=1, min_row=2, max_row=11))
bar3.height, bar3.width, bar3.legend = 8, 16, None
db.add_chart(bar3, "J25")

wb.save("spend_dashboard.xlsx")
print("saved")
